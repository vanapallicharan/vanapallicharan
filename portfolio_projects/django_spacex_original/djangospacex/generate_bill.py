import sqlite3
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime

# Update this path to the location where you uploaded the template
template_path = r'C:\Users\chara\OneDrive\Desktop\WORK\template.xlsx'

def parse_products(input_string):
    pattern = r'(\d+X\d+)'
    matches = re.findall(pattern, input_string)
    products = []

    for match in matches:
        s_no, qty = map(int, match.split('X'))
        products.append((s_no, qty))

    return products

def fetch_product_details(s_no):
    conn = sqlite3.connect('products.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM Products WHERE S_NO=?', (s_no,))
    product = cursor.fetchone()
    conn.close()

    if product:
        return product[1], product[2]  # Description and Price
    else:
        return None, None

def calculate_prices(price, qty, discount=0):
    price_after_discount = price * (1 - discount / 100)
    total_amount = price_after_discount * qty
    return price_after_discount, total_amount

def generate_unique_filename(base_path, customer_name):
    sanitized_name = re.sub(r'[^a-zA-Z0-9]', '_', customer_name)  # Sanitize customer name for filename
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{sanitized_name}_invoice_{timestamp}.xlsx"
    return os.path.join(base_path, filename)

def copy_cell_style(source_cell, target_cell):
    target_cell.font = Font(bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            underline=source_cell.font.underline,
                            color=source_cell.font.color,
                            size=source_cell.font.size,
                            name=source_cell.font.name)
    
    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      wrap_text=source_cell.alignment.wrap_text,
                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                      indent=source_cell.alignment.indent,
                                      text_rotation=source_cell.alignment.text_rotation)

def extract_initial_styles(ws):
    styles = {}
    for row in ws.iter_rows():
        for cell in row:
            styles[cell.coordinate] = {
                'font': cell.font.copy(),
                'alignment': cell.alignment.copy(),
                'border': cell.border.copy(),
                'fill': cell.fill.copy(),
                'number_format': cell.number_format,
                'protection': cell.protection.copy()
            }
    return styles

def apply_styles(ws, styles):
    for coord, style in styles.items():
        cell = ws[coord]
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
        cell.fill = style['fill']
        cell.number_format = style['number_format']
        cell.protection = style['protection']

def unmerge_cells(ws):
    merged_cells = list(ws.merged_cells.ranges)  # Make a copy of the set
    for merged_range in merged_cells:
        ws.unmerge_cells(str(merged_range))

def remerge_cells(ws):
    # Re-merge specific cells based on the template structure.
    ws.merge_cells('A2:K2')
    ws.merge_cells('A4:F4')
    ws.merge_cells('J4:K4')
    ws.merge_cells('A5:F6')
    ws.merge_cells('A7:F7')
    ws.merge_cells('J5:K5')
    ws.merge_cells('J6:K6')
    ws.merge_cells('J7:K7')
    ws.merge_cells('A9:E9')
    ws.merge_cells('F9:K9')
    ws.merge_cells('A10:E10')
    ws.merge_cells('F10:K10')
    ws.merge_cells('A11:E11')
    ws.merge_cells('F11:K12')
    ws.merge_cells('A12:E12')
    ws.merge_cells('B14:F14')
    ws.merge_cells('J14:K14')
    for row in range(15, 28):
        ws.merge_cells(f'B{row}:F{row}')
        ws.merge_cells(f'J{row}:K{row}')
    ws.merge_cells('E25:I25')
    ws.merge_cells('E27:I27')
    for row in range(29, 36):
        ws.merge_cells(f'A{row}:K{row}')
    ws.merge_cells('A37:K37')
    ws.merge_cells('A38:K38')
    ws.merge_cells('A40:K40')
    ws.merge_cells('A44:K44')

def get_top_left_cell(ws, cell):
    """ Returns the top-left cell if the cell is part of a merged range, otherwise returns the cell itself. """
    for merged_range in list(ws.merged_cells.ranges):  # Make a copy of the set
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

def set_cell_value_safe(ws, cell, value):
    """ Sets the cell value safely, handling merged cells. """
    top_left_cell = get_top_left_cell(ws, cell)
    if isinstance(top_left_cell, openpyxl.cell.MergedCell):
        # Unmerge the cell before setting value
        ws.unmerge_cells(str(top_left_cell))

        # Set the value in the top-left cell of the merged range
        top_left_cell.value = value

        # Re-merge the cells after setting value
        ws.merge_cells(str(top_left_cell))
    else:
        cell.value = value

def generate_bill(product_data, discount, name, address, base_path):
    unique_filename = generate_unique_filename(base_path, name)

    # Load the workbook and extract initial styles
    wb = load_workbook(template_path)
    ws = wb.active
    initial_styles = extract_initial_styles(ws)

    # Unmerge cells before setting values
    unmerge_cells(ws)

    try:
        # Update header information
        ws['A10'].value = name
        ws['A10'].alignment = Alignment(horizontal='left')
        ws['A12'].value = address
        ws['A12'].alignment = Alignment(horizontal='left')
        ws['J4'].value = datetime.now().strftime("%Y-%m-%d")
        ws['J4'].alignment = Alignment(horizontal='center')

        # Add column headers
        ws['A14'].value = 'S. No'
        ws['B14'].value = 'Description'
        ws['G14'].value = 'QTY'
        ws['H14'].value = 'Price'
        ws['I14'].value = 'Price After Discount'
        ws['J14'].value = 'Amount'

        # Start adding product details from row 15
        row_start = 15
        product_count = 0
        for index, product in enumerate(product_data):
            if row_start + product_count > 24:
                # Insert a new row before 25th row without disturbing subsequent rows
                ws.insert_rows(25)
                ws.merge_cells(f'B{25}:F{25}')
                ws.merge_cells(f'J{25}:K{25}')
                ws.move_range(f'A{25}:K{ws.max_row}', rows=1)

            s_no, qty = product
            description, price = fetch_product_details(s_no)
            if description and price is not None:
                price_after_discount, total_amount = calculate_prices(price, qty, discount)
                ws.cell(row=row_start + product_count, column=1, value=s_no)
                ws.cell(row=row_start + product_count, column=2, value=description)
                ws.cell(row=row_start + product_count, column=7, value=qty)
                ws.cell(row=row_start + product_count, column=8, value=price)
                ws.cell(row=row_start + product_count, column=9, value=price_after_discount)
                ws.cell(row=row_start + product_count, column=10, value=total_amount)
                product_count += 1
            else:
                print(f"No details found for S_NO: {s_no}")

        # Reapply initial styles
        apply_styles(ws, initial_styles)

        # Re-merge cells after setting values
        remerge_cells(ws)

        # Save the finalized bill
        wb.save(unique_filename)
        print(f"Bill generated and saved as '{unique_filename}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

def process_and_generate_bill(input_string, discount, name, address, base_path):
    products = parse_products(input_string)
    generate_bill(products, discount, name, address, base_path)

if __name__ == '__main__':
    input_string = input("Enter the product details (e.g., *12X1,100X5,56X2#): ")
    discount = float(input("Enter the discount percentage (e.g., 30): "))
    name = input("Enter the customer name: ")
    address = input("Enter the customer address: ")
    base_path = r'C:\Users\chara\OneDrive\Desktop\WORK\bills'


    try:
        process_and_generate_bill(input_string, discount, name, address, base_path)
    except Exception as e:
        print(f"An error occurred: {e}")
